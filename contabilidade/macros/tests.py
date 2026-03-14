import json
from io import BytesIO
from unittest.mock import patch
import csv as csv_reader

from django.contrib.auth.models import User
from django.core.cache import cache
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from datetime import timedelta

from contabilidade.macros.models import MacroLead, MacroRun
from contabilidade.macros.services import upsert_rows


class MacroServicesTests(TestCase):
    def test_upsert_inserts_duplicate_rows_with_same_store_id(self):
        first = {
            "ID da loja": "1001",
            "Cidade": "Sao Paulo",
            "Nome do estabelecimento": "Loja Teste",
            "Telefone do representante do estabelecimento": "(11) 99999-0000",
            "Status do contrato": "Ativo",
        }
        second = {
            "ID da loja": "1001",
            "Cidade": "Sao Paulo",
            "Nome do estabelecimento": "Loja Teste",
            "Telefone do representante do estabelecimento": "11999990000",
            "Status do contrato": "Pendente",
        }
        upsert_rows([first], default_source="api")
        result = upsert_rows([second], default_source="api")
        self.assertEqual(result["created"], 1)
        self.assertEqual(result["updated"], 0)
        self.assertEqual(MacroLead.objects.count(), 2)
        self.assertEqual(
            list(
                MacroLead.objects.filter(store_id="1001")
                .values_list("contract_status", flat=True)
                .order_by("id")
            ),
            ["Ativo", "Pendente"],
        )

    def test_upsert_parses_lead_created_at(self):
        row = {
            "ID da loja": "1002",
            "ID do signatario": "9002",
            "Cidade": "Sao Paulo",
            "Nome do estabelecimento": "Loja Data",
            "Telefone do representante do estabelecimento": "11999990000",
            "Horario de criacao do lead": "2026-02-02 13:45:20 UTC-3",
            "Seu Negocio na 99": "Nao ativado",
        }
        upsert_rows([row], default_source="api")
        lead = MacroLead.objects.first()
        self.assertIsNotNone(lead.lead_created_at)
        self.assertEqual(lead.business_99_status, "Nao ativado")
        self.assertEqual(lead.store_id, "1002")
        self.assertEqual(lead.signatory_id, "9002")

    def test_upsert_trims_oversized_values(self):
        row = {
            "ID da loja": "1003",
            "Cidade": "S" * 400,
            "Nome do estabelecimento": "L" * 400,
            "Status do contrato": "A" * 150,
            "Telefone do representante do estabelecimento": "9" * 80,
            "Categoria da empresa": "C" * 400,
            "Endereco": "Rua X",
        }
        result = upsert_rows([row], default_source="api")
        self.assertEqual(result["created"], 1)
        lead = MacroLead.objects.first()
        self.assertEqual(len(lead.city), 255)
        self.assertEqual(len(lead.establishment_name), 255)
        self.assertEqual(len(lead.contract_status), 100)
        self.assertEqual(len(lead.representative_phone), 50)
        self.assertEqual(len(lead.company_category), 255)

    def test_upsert_always_creates_new_rows(self):
        row = {
            "ID da loja": "1004",
            "Cidade": "Sao Paulo",
            "Nome do estabelecimento": "Loja Corrida",
            "Telefone do representante do estabelecimento": "11999990000",
            "Status do contrato": "Ativo",
        }
        first = upsert_rows([row], default_source="api")
        second = upsert_rows([row], default_source="api")
        self.assertEqual(first["created"], 1)
        self.assertEqual(second["created"], 1)
        self.assertEqual(MacroLead.objects.count(), 2)


@override_settings(
    MACRO_API_TOKEN="token123",
    MACRO_API_ALLOWED_IPS=[],
    MACRO_API_RATE_LIMIT_PER_MINUTE=100,
)
class MacroApiImportTests(TestCase):
    def setUp(self):
        cache.clear()
        self.client = Client()
        self.url = reverse("macro_api_import")

    def test_api_import_with_token(self):
        payload = [{"Cidade": "Rio", "Nome do estabelecimento": "Loja API"}]
        resp = self.client.post(
            self.url,
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_AUTHORIZATION="Bearer token123",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(MacroLead.objects.filter(establishment_name="Loja API").count(), 1)

    def test_api_import_marks_error_if_processing_fails(self):
        payload = [{"Cidade": "Rio", "Nome do estabelecimento": "Loja API"}]
        with patch("contabilidade.macros.views.upsert_rows", side_effect=RuntimeError("boom")):
            resp = self.client.post(
                self.url,
                data=json.dumps(payload),
                content_type="application/json",
                HTTP_AUTHORIZATION="Bearer token123",
            )
        self.assertEqual(resp.status_code, 500)
        run = MacroRun.objects.first()
        self.assertIsNotNone(run)
        self.assertEqual(run.status, "error")

    @override_settings(MACRO_API_RATE_LIMIT_PER_MINUTE=1)
    def test_api_rate_limit_blocks_second_request(self):
        payload = [{"Cidade": "Rio", "Nome do estabelecimento": "Loja 1"}]
        first = self.client.post(
            self.url,
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_AUTHORIZATION="Bearer token123",
        )
        second = self.client.post(
            self.url,
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_AUTHORIZATION="Bearer token123",
        )
        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 429)

    def test_api_import_saves_meta_pages_and_collected_total(self):
        payload = {
            "rows": [{"Cidade": "Recife", "Nome do estabelecimento": "Loja Meta"}],
            "meta": {
                "execution_id": "exec-meta-1",
                "pages_processed": 62,
                "collected_total": 3072,
                "deduplicated_total": 1800,
                "batch_index": 1,
                "batch_total": 1,
                "sent_after": 400,
            },
        }
        resp = self.client.post(
            self.url,
            data=json.dumps(payload),
            content_type="application/json",
            HTTP_AUTHORIZATION="Bearer token123",
        )
        self.assertEqual(resp.status_code, 200)
        run = MacroRun.objects.first()
        self.assertIsNotNone(run)
        self.assertEqual(run.pages_processed, 62)
        self.assertEqual(run.total_collected, 3072)
        self.assertEqual(run.total_deduplicated, 1800)
        self.assertEqual(run.total_sent, 400)
        self.assertEqual(run.status, "success")
        self.assertIn("Importacao API concluida.", run.message)

    def test_api_import_consolidates_batches_by_execution_id(self):
        headers = {
            "content_type": "application/json",
            "HTTP_AUTHORIZATION": "Bearer token123",
        }
        first_payload = {
            "rows": [{"Cidade": "Recife", "Nome do estabelecimento": "Loja Lote 1"}],
            "meta": {
                "execution_id": "exec-batch-1",
                "batch_index": 1,
                "batch_total": 2,
                "pages_processed": 18,
                "collected_total": 2625,
                "deduplicated_total": 475,
                "sent_after": 400,
            },
        }
        second_payload = {
            "rows": [{"Cidade": "Recife", "Nome do estabelecimento": "Loja Lote 2"}],
            "meta": {
                "execution_id": "exec-batch-1",
                "batch_index": 2,
                "batch_total": 2,
                "pages_processed": 18,
                "collected_total": 2625,
                "deduplicated_total": 475,
                "sent_after": 475,
            },
        }

        first_resp = self.client.post(self.url, data=json.dumps(first_payload), **headers)
        second_resp = self.client.post(self.url, data=json.dumps(second_payload), **headers)

        self.assertEqual(first_resp.status_code, 200)
        self.assertEqual(second_resp.status_code, 200)

        runs = MacroRun.objects.filter(run_type="api", execution_id="exec-batch-1")
        self.assertEqual(runs.count(), 1)
        run = runs.first()
        self.assertEqual(run.status, "success")
        self.assertEqual(run.total_collected, 2625)
        self.assertEqual(run.total_deduplicated, 475)
        self.assertEqual(run.total_sent, 475)
        self.assertEqual(run.pages_processed, 18)
        self.assertIn("Lote 2/2", run.message)


class MacroScreenTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="staff", password="123456", is_staff=True)
        self.client = Client()
        self.client.login(username="staff", password="123456")

    def test_macro_page_requires_staff(self):
        resp = self.client.get(reverse("macro_list"))
        self.assertEqual(resp.status_code, 200)
        collect_resp = self.client.get(reverse("macro_collect"))
        self.assertEqual(collect_resp.status_code, 200)

    def test_macro_pages_open_when_optional_columns_are_unavailable(self):
        lead_columns = {
            "id",
            "source",
            "city",
            "target_region",
            "establishment_name",
            "representative_name",
            "contract_status",
            "representative_phone",
            "company_category",
            "address",
            "unique_key",
            "first_seen_at",
            "last_seen_at",
        }
        run_columns = {
            "id",
            "run_type",
            "status",
            "source",
            "started_at",
            "finished_at",
            "request_ip",
            "total_collected",
            "total_received",
            "total_sent",
            "created_count",
            "updated_count",
            "ignored_count",
            "invalid_count",
            "message",
            "triggered_by_id",
        }

        with patch("contabilidade.macros.views._macrolead_db_columns", return_value=lead_columns), patch(
            "contabilidade.macros.views._macrorun_db_columns", return_value=run_columns
        ):
            resp = self.client.get(reverse("macro_list"))
            self.assertEqual(resp.status_code, 200)
            self.assertFalse(resp.context["store_id_enabled"])
            self.assertFalse(resp.context["signatory_id_enabled"])
            self.assertFalse(resp.context["export_tracking_enabled"])
            self.assertFalse(resp.context["business_99_enabled"])
            self.assertFalse(resp.context["lead_created_at_enabled"])
            self.assertFalse(resp.context["run_pages_enabled"])

            collect_resp = self.client.get(reverse("macro_collect"))
            self.assertEqual(collect_resp.status_code, 200)
            self.assertFalse(collect_resp.context["run_pages_enabled"])

    def test_macro_pages_open_when_run_table_is_unavailable(self):
        with patch("contabilidade.macros.views._macrorun_db_columns", return_value=set()):
            resp = self.client.get(reverse("macro_list"))
            self.assertEqual(resp.status_code, 200)
            self.assertFalse(resp.context["run_table_ready"])

            collect_resp = self.client.get(reverse("macro_collect"))
            self.assertEqual(collect_resp.status_code, 200)

    def test_export_xlsx_works(self):
        MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="Centro",
            establishment_name="Loja X",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999998888",
            representative_phone_norm="5521999998888",
            company_category="Restaurante",
            address="Rua Teste",
            unique_key="k1",
        )
        resp = self.client.get(reverse("macro_export_xlsx"))
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resp["Content-Type"],
        )

    def test_export_csv_with_selected_columns_and_limit(self):
        MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="Centro",
            establishment_name="Loja X",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999998888",
            representative_phone_norm="5521999998888",
            company_category="Restaurante",
            address="Rua Teste",
            unique_key="csv-k1",
        )
        MacroLead.objects.create(
            source="api",
            city="Sao Paulo",
            target_region="Sul",
            establishment_name="Loja Y",
            representative_name="Bia",
            contract_status="Pendente",
            representative_phone="11999998888",
            representative_phone_norm="5511999998888",
            company_category="Pizza",
            address="Rua Teste 2",
            unique_key="csv-k2",
        )
        resp = self.client.get(
            reverse("macro_export_csv"),
            data={
                "export_limit": "1",
                "export_fields": ["city", "establishment_name", "representative_phone"],
            },
        )
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode("utf-8")
        rows = list(csv_reader.reader(body.splitlines()))
        self.assertEqual(rows[0], ["Cidade", "Nome do estabelecimento", "Telefone do representante do estabelecimento"])
        self.assertEqual(len(rows), 2)

    def test_export_xlsx_with_selected_columns_and_limit(self):
        from openpyxl import load_workbook

        MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="Centro",
            establishment_name="Loja X",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999998888",
            representative_phone_norm="5521999998888",
            company_category="Restaurante",
            address="Rua Teste",
            unique_key="xlsx-k1",
        )
        MacroLead.objects.create(
            source="api",
            city="Sao Paulo",
            target_region="Sul",
            establishment_name="Loja Y",
            representative_name="Bia",
            contract_status="Pendente",
            representative_phone="11999998888",
            representative_phone_norm="5511999998888",
            company_category="Pizza",
            address="Rua Teste 2",
            unique_key="xlsx-k2",
        )

        resp = self.client.get(
            reverse("macro_export_xlsx"),
            data={
                "export_limit": "1",
                "export_fields": ["city", "source", "last_seen_at"],
            },
        )
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, ["Cidade", "Fonte", "Ultima captura"])
        self.assertEqual(ws.max_row, 2)

    def test_export_marks_leads_and_export_status_filter(self):
        lead_exported = MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="Centro",
            establishment_name="Loja Exportada",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999998888",
            representative_phone_norm="5521999998888",
            company_category="Restaurante",
            address="Rua Teste",
            unique_key="exp-mark-1",
        )
        lead_pending = MacroLead.objects.create(
            source="api",
            city="Sao Paulo",
            target_region="Sul",
            establishment_name="Loja Pendente",
            representative_name="Bia",
            contract_status="Pendente",
            representative_phone="11999998888",
            representative_phone_norm="5511999998888",
            company_category="Pizza",
            address="Rua Teste 2",
            unique_key="exp-mark-2",
        )

        export_resp = self.client.get(
            reverse("macro_export_csv"),
            data={
                "q": "Loja Exportada",
                "mark_exported": "1",
                "export_fields": ["establishment_name"],
            },
        )
        self.assertEqual(export_resp.status_code, 200)

        lead_exported.refresh_from_db()
        lead_pending.refresh_from_db()
        self.assertIsNotNone(lead_exported.exported_at)
        self.assertTrue(lead_exported.export_batch_id)
        self.assertIsNone(lead_pending.exported_at)
        self.assertEqual(lead_pending.export_batch_id, "")

        exported_resp = self.client.get(reverse("macro_list"), data={"export_status": "exported"})
        exported_names = {
            row.establishment_name for row in exported_resp.context["page_obj"].object_list
        }
        self.assertIn("Loja Exportada", exported_names)
        self.assertNotIn("Loja Pendente", exported_names)

        not_exported_resp = self.client.get(
            reverse("macro_list"),
            data={"export_status": "not_exported"},
        )
        not_exported_names = {
            row.establishment_name for row in not_exported_resp.context["page_obj"].object_list
        }
        self.assertIn("Loja Pendente", not_exported_names)
        self.assertNotIn("Loja Exportada", not_exported_names)

    def test_export_without_mark_keeps_export_fields_empty(self):
        lead = MacroLead.objects.create(
            source="api",
            city="Recife",
            target_region="Norte",
            establishment_name="Loja Sem Marca",
            representative_name="Carlos",
            contract_status="Ativo",
            representative_phone="81999998888",
            representative_phone_norm="5581999998888",
            company_category="Lanches",
            address="Rua Teste 3",
            unique_key="exp-mark-3",
        )
        resp = self.client.get(
            reverse("macro_export_csv"),
            data={"q": "Loja Sem Marca", "export_fields": ["establishment_name"]},
        )
        self.assertEqual(resp.status_code, 200)
        lead.refresh_from_db()
        self.assertIsNone(lead.exported_at)
        self.assertEqual(lead.export_batch_id, "")

    def test_download_local_agent_files(self):
        exe_resp = self.client.get(reverse("macro_download_local_agent_exe"))
        mac_resp = self.client.get(reverse("macro_download_local_agent_mac"))
        py_resp = self.client.get(reverse("macro_download_local_agent_py"))
        bat_resp = self.client.get(reverse("macro_download_local_agent_bat"))
        self.assertIn(exe_resp.status_code, (200, 404))
        self.assertEqual(mac_resp.status_code, 200)
        self.assertIn("attachment; filename=", mac_resp["Content-Disposition"])
        self.assertEqual(py_resp.status_code, 200)
        self.assertEqual(bat_resp.status_code, 200)
        self.assertIn("attachment; filename=", py_resp["Content-Disposition"])
        self.assertIn("attachment; filename=", bat_resp["Content-Disposition"])

    def test_delete_filtered_leads(self):
        MacroLead.objects.create(
            source="api",
            city="Rio de Janeiro",
            target_region="R1",
            establishment_name="Loja RJ",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999998888",
            representative_phone_norm="5521999998888",
            company_category="Brasileira",
            address="Rua A",
            unique_key="del-1",
        )
        MacroLead.objects.create(
            source="api",
            city="Sao Paulo",
            target_region="R2",
            establishment_name="Loja SP",
            representative_name="Joao",
            contract_status="Ativo",
            representative_phone="11999998888",
            representative_phone_norm="5511999998888",
            company_category="Pizza",
            address="Rua B",
            unique_key="del-2",
        )
        resp = self.client.post(
            reverse("macro_delete_filtered"),
            data={"city": "Rio de Janeiro", "confirm_text": "EXCLUIR"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(MacroLead.objects.count(), 1)
        self.assertEqual(MacroLead.objects.first().city, "Sao Paulo")

    def test_delete_all_and_runs(self):
        MacroLead.objects.create(
            source="api",
            city="Campinas",
            target_region="R3",
            establishment_name="Loja C",
            representative_name="Maria",
            contract_status="Pendente",
            representative_phone="19999998888",
            representative_phone_norm="5519999998888",
            company_category="Lanches",
            address="Rua C",
            unique_key="del-3",
        )
        MacroRun.objects.create(
            run_type="csv",
            status="success",
            source="csv",
            message="ok",
        )
        run_resp = self.client.post(
            reverse("macro_delete_runs"),
            data={"confirm_text": "LIMPAR HISTORICO", "next": "collect"},
        )
        all_resp = self.client.post(
            reverse("macro_delete_all"),
            data={"confirm_text": "APAGAR TUDO", "next": "collect"},
        )
        self.assertEqual(run_resp.status_code, 302)
        self.assertEqual(all_resp.status_code, 302)
        self.assertEqual(MacroRun.objects.count(), 0)
        self.assertEqual(MacroLead.objects.count(), 0)

    def test_delete_specific_source(self):
        MacroLead.objects.create(
            source="csv",
            city="Rio",
            target_region="R1",
            establishment_name="Loja CSV",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999990000",
            representative_phone_norm="5521999990000",
            company_category="Brasileira",
            address="Rua 1",
            unique_key="src-1",
        )
        MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="R2",
            establishment_name="Loja API",
            representative_name="Joao",
            contract_status="Ativo",
            representative_phone="21999991111",
            representative_phone_norm="5521999991111",
            company_category="Pizza",
            address="Rua 2",
            unique_key="src-2",
        )
        resp = self.client.post(
            reverse("macro_delete_source"),
            data={"source": "csv", "confirm_text": "EXCLUIR BASE", "next": "database"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(MacroLead.objects.count(), 1)
        self.assertEqual(MacroLead.objects.first().source, "api")

    def test_delete_specific_run_item(self):
        run = MacroRun.objects.create(
            run_type="csv",
            status="success",
            source="csv",
            message="ok",
        )
        resp = self.client.post(
            reverse("macro_delete_run_item", args=[run.id]),
            data={"next": "collect"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(MacroRun.objects.filter(id=run.id).exists())

    def test_block_and_unblock_phone_updates_same_number(self):
        lead_a = MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="R1",
            establishment_name="Loja A",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999990000",
            representative_phone_norm="5521999990000",
            company_category="Brasileira",
            address="Rua 1",
            unique_key="block-1",
        )
        lead_b = MacroLead.objects.create(
            source="api",
            city="Niteroi",
            target_region="R2",
            establishment_name="Loja B",
            representative_name="Bia",
            contract_status="Ativo",
            representative_phone="(21) 99999-0000",
            representative_phone_norm="5521999990000",
            company_category="Pizza",
            address="Rua 2",
            unique_key="block-2",
        )

        block_resp = self.client.post(reverse("macro_block_phone", args=[lead_a.id]), data={})
        self.assertEqual(block_resp.status_code, 302)
        self.assertEqual(MacroLead.objects.filter(is_blocked_number=True).count(), 2)

        unblock_resp = self.client.post(reverse("macro_unblock_phone", args=[lead_b.id]), data={})
        self.assertEqual(unblock_resp.status_code, 302)
        self.assertEqual(MacroLead.objects.filter(is_blocked_number=True).count(), 0)

    def test_delete_blocked_respects_filter(self):
        MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="R1",
            establishment_name="Loja Bloq",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999990000",
            representative_phone_norm="5521999990000",
            is_blocked_number=True,
            company_category="Brasileira",
            address="Rua 1",
            unique_key="blocked-del-1",
        )
        MacroLead.objects.create(
            source="api",
            city="Sao Paulo",
            target_region="R2",
            establishment_name="Loja Livre",
            representative_name="Joao",
            contract_status="Ativo",
            representative_phone="11999998888",
            representative_phone_norm="5511999998888",
            is_blocked_number=False,
            company_category="Pizza",
            address="Rua 2",
            unique_key="blocked-del-2",
        )
        resp = self.client.post(
            reverse("macro_delete_blocked"),
            data={"blocked": "yes", "confirm_text": "EXCLUIR BLOQUEADOS"},
        )
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(MacroLead.objects.count(), 1)
        self.assertEqual(MacroLead.objects.first().city, "Sao Paulo")

    def test_filter_duplicate_phones(self):
        MacroLead.objects.create(
            source="api",
            store_id="dup-001",
            city="Rio",
            target_region="R1",
            establishment_name="Loja Dup 1",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999990000",
            representative_phone_norm="5521999990000",
            company_category="Brasileira",
            address="Rua 1",
            unique_key="dup-filter-1",
        )
        MacroLead.objects.create(
            source="api",
            store_id="dup-001",
            city="Niteroi",
            target_region="R2",
            establishment_name="Loja Dup 2",
            representative_name="Bia",
            contract_status="Ativo",
            representative_phone="(21) 99999-0000",
            representative_phone_norm="5521999990000",
            company_category="Pizza",
            address="Rua 2",
            unique_key="dup-filter-2",
        )
        MacroLead.objects.create(
            source="api",
            store_id="uniq-001",
            city="Sao Paulo",
            target_region="R3",
            establishment_name="Loja Unica",
            representative_name="Joao",
            contract_status="Ativo",
            representative_phone="11999998888",
            representative_phone_norm="5511999998888",
            company_category="Lanches",
            address="Rua 3",
            unique_key="dup-filter-3",
        )

        duplicates_resp = self.client.get(reverse("macro_list"), data={"phone_dup": "duplicates"})
        duplicates_page = list(duplicates_resp.context["page_obj"].object_list)
        self.assertEqual(duplicates_resp.status_code, 200)
        self.assertEqual(len(duplicates_page), 2)

        unique_resp = self.client.get(reverse("macro_list"), data={"phone_dup": "unique"})
        unique_page = list(unique_resp.context["page_obj"].object_list)
        self.assertEqual(unique_resp.status_code, 200)
        self.assertEqual(len(unique_page), 1)
        self.assertEqual(unique_page[0].city, "Sao Paulo")

    def test_filter_by_representative_presence(self):
        MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="R1",
            establishment_name="Loja Com Rep",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999990000",
            representative_phone_norm="5521999990000",
            company_category="Brasileira",
            address="Rua 1",
            unique_key="rep-filter-1",
        )
        MacroLead.objects.create(
            source="api",
            city="Sao Paulo",
            target_region="R2",
            establishment_name="Loja Sem Rep",
            representative_name="",
            contract_status="Ativo",
            representative_phone="11999998888",
            representative_phone_norm="5511999998888",
            company_category="Pizza",
            address="Rua 2",
            unique_key="rep-filter-2",
        )
        MacroLead.objects.create(
            source="api",
            city="Curitiba",
            target_region="R3",
            establishment_name="Loja Sem Rep Traco",
            representative_name="-",
            contract_status="Ativo",
            representative_phone="41999998888",
            representative_phone_norm="5541999998888",
            company_category="Lanches",
            address="Rua 3",
            unique_key="rep-filter-3",
        )

        with_resp = self.client.get(reverse("macro_list"), data={"representative_presence": "with"})
        with_page = list(with_resp.context["page_obj"].object_list)
        self.assertEqual(with_resp.status_code, 200)
        self.assertEqual(len(with_page), 1)
        self.assertEqual(with_page[0].establishment_name, "Loja Com Rep")

        without_resp = self.client.get(reverse("macro_list"), data={"representative_presence": "without"})
        without_page = list(without_resp.context["page_obj"].object_list)
        self.assertEqual(without_resp.status_code, 200)
        self.assertEqual(len(without_page), 2)
        names = {item.establishment_name for item in without_page}
        self.assertIn("Loja Sem Rep", names)
        self.assertIn("Loja Sem Rep Traco", names)

    def test_export_csv_respects_representative_presence_filter(self):
        MacroLead.objects.create(
            source="api",
            city="Rio",
            target_region="R1",
            establishment_name="Loja Com Rep Export",
            representative_name="Ana",
            contract_status="Ativo",
            representative_phone="21999990000",
            representative_phone_norm="5521999990000",
            company_category="Brasileira",
            address="Rua 1",
            unique_key="rep-export-1",
        )
        MacroLead.objects.create(
            source="api",
            city="Sao Paulo",
            target_region="R2",
            establishment_name="Loja Sem Rep Export",
            representative_name="",
            contract_status="Ativo",
            representative_phone="11999998888",
            representative_phone_norm="5511999998888",
            company_category="Pizza",
            address="Rua 2",
            unique_key="rep-export-2",
        )

        resp = self.client.get(
            reverse("macro_export_csv"),
            data={
                "representative_presence": "without",
                "export_fields": ["establishment_name", "representative_name"],
            },
        )
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode("utf-8")
        self.assertIn("Loja Sem Rep Export", body)
        self.assertNotIn("Loja Com Rep Export", body)

    @override_settings(MACRO_RUN_STALE_MINUTES=1)
    def test_collect_page_closes_stale_running_runs(self):
        run = MacroRun.objects.create(
            run_type="api",
            status="running",
            source="api",
            message="aguardando",
        )
        MacroRun.objects.filter(id=run.id).update(started_at=timezone.now() - timedelta(minutes=10))
        resp = self.client.get(reverse("macro_collect"))
        self.assertEqual(resp.status_code, 200)
        run.refresh_from_db()
        self.assertEqual(run.status, "error")
        self.assertIsNotNone(run.finished_at)
